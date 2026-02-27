#!/usr/bin/env python3
"""
账单截图解析 - Excel 导出脚本

将解析的账单数据导出为 Excel 文件
"""

import argparse
import json
import os
import sys
from datetime import datetime

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("请先安装 openpyxl: pip install openpyxl")
    sys.exit(1)


# Excel 表头
HEADERS = ["日期", "类型", "金额", "类别", "描述", "平台", "订单号"]

# 样式定义
HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center")
BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

# 类型映射
TYPE_MAP = {
    "expense": "支出",
    "income": "收入",
}

# 平台映射
PLATFORM_MAP = {
    "alipay": "支付宝",
    "wechat": "微信",
}

# 过滤关键词（默认排除）
FILTER_KEYWORDS = [
    "余额宝", "收益发放", "基金分红", "理财",
    "您于", "商户单号"
]


def should_filter(data):
    """判断是否应该过滤此记录"""
    description = data.get("description", "")
    for keyword in FILTER_KEYWORDS:
        if keyword in description:
            return True
    return False


def create_workbook():
    """创建新的工作簿"""
    wb = Workbook()
    ws = wb.active
    ws.title = "账单记录"

    # 设置表头
    for col, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = BORDER

    # 设置列宽
    column_widths = [20, 8, 12, 10, 40, 10, 30]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    return wb, ws


def append_row(ws, row_num, data):
    """添加一行数据"""
    row_data = [
        data.get("date", ""),
        TYPE_MAP.get(data.get("type", ""), data.get("type", "")),
        data.get("amount", 0),
        data.get("category", ""),
        data.get("description", ""),
        PLATFORM_MAP.get(data.get("platform", ""), data.get("platform", "")),
        data.get("orderId", ""),
    ]

    for col, value in enumerate(row_data, 1):
        cell = ws.cell(row=row_num, column=col, value=value)
        cell.border = BORDER
        cell.alignment = Alignment(vertical="center")

        # 金额列右对齐
        if col == 3:
            cell.alignment = Alignment(horizontal="right", vertical="center")
            cell.number_format = "#,##0.00"

        # 类型列颜色
        if col == 2:
            if value == "支出":
                cell.font = Font(color="FF0000")
            elif value == "收入":
                cell.font = Font(color="00AA00")


def get_existing_order_ids(ws):
    """获取已有的订单号列表"""
    order_ids = set()
    for row in ws.iter_rows(min_row=2, max_col=7, values_only=True):
        if row[6]:  # 订单号在第7列
            order_ids.add(str(row[6]))
    return order_ids


def export_to_excel(data_list, output_path, append_mode=False, no_filter=False):
    """
    导出数据到 Excel

    Args:
        data_list: 数据列表
        output_path: 输出文件路径
        append_mode: 是否追加模式
        no_filter: 是否禁用过滤（保留所有记录）
    """
    if not data_list:
        print("没有数据需要导出")
        return False

    if append_mode and os.path.exists(output_path):
        wb = load_workbook(output_path)
        ws = wb.active
        existing_order_ids = get_existing_order_ids(ws)
        start_row = ws.max_row + 1
    else:
        wb, ws = create_workbook()
        existing_order_ids = set()
        start_row = 2

    # 统计
    added_count = 0
    skipped_count = 0
    filtered_count = 0

    for data in data_list:
        # 过滤检查
        if not no_filter and should_filter(data):
            filtered_count += 1
            continue

        order_id = str(data.get("orderId", ""))
        if order_id and order_id in existing_order_ids:
            skipped_count += 1
            continue

        append_row(ws, start_row, data)
        start_row += 1
        added_count += 1

    # 保存文件
    wb.save(output_path)

    print(f"导出完成: {output_path}")
    print(f"  新增: {added_count} 条")
    if filtered_count > 0:
        print(f"  已过滤: {filtered_count} 条")
    if skipped_count > 0:
        print(f"  跳过重复: {skipped_count} 条")

    return True


def main():
    parser = argparse.ArgumentParser(description="导出账单数据到 Excel")
    parser.add_argument("--data", required=True, help="JSON 格式的账单数据")
    parser.add_argument("--output", required=True, help="输出文件路径")
    parser.add_argument("--append", action="store_true", help="追加到已有文件")
    parser.add_argument("--no-filter", action="store_true", help="禁用过滤，保留所有记录")

    args = parser.parse_args()

    try:
        data_list = json.loads(args.data)
        if not isinstance(data_list, list):
            data_list = [data_list]
    except json.JSONDecodeError as e:
        print(f"JSON 解析错误: {e}")
        sys.exit(1)

    export_to_excel(data_list, args.output, args.append, args.no_filter)


if __name__ == "__main__":
    main()
