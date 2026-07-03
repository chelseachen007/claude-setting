#!/usr/bin/env python3
"""
账单文件解析器
支持：支付宝 CSV（GBK 编码）、微信 XLSX
输出：按月分组的 JSON，供 process_and_write.py 使用
"""

import argparse
import csv
import json
import os
import re
import sys
from datetime import datetime

try:
    import openpyxl
except ImportError:
    openpyxl = None


def detect_file_type(filepath):
    """自动检测文件类型：'alipay_csv' | 'wechat_xlsx' | None"""
    ext = os.path.splitext(filepath)[1].lower()

    if ext == '.csv':
        # 尝试读取前几行确认是支付宝
        for enc in ('gbk', 'utf-8', 'gb2312'):
            try:
                with open(filepath, 'r', encoding=enc) as f:
                    head = ''.join(f.readline() for _ in range(20))
                if '支付宝' in head or '交易时间,交易分类' in head:
                    return 'alipay_csv'
                break
            except (UnicodeDecodeError, UnicodeError):
                continue
        return 'alipay_csv'  # CSV 默认当支付宝处理

    if ext == '.xlsx':
        if openpyxl is None:
            print("错误：解析 xlsx 需要 openpyxl，请先安装：pip3 install openpyxl", file=sys.stderr)
            sys.exit(1)
        wb = openpyxl.load_workbook(filepath, read_only=True)
        ws = wb.active
        head = ''
        for row in ws.iter_rows(min_row=1, max_row=20, values_only=True):
            for cell in row:
                if cell:
                    head += str(cell) + ' '
        wb.close()
        if '微信' in head:
            return 'wechat_xlsx'
        return 'wechat_xlsx'  # XLSX 默认当微信处理

    return None


def parse_alipay_csv(filepath, target_months=None):
    """解析支付宝 CSV，返回记录列表"""
    records = []
    encoding = None

    for enc in ('gbk', 'utf-8', 'gb2312'):
        try:
            with open(filepath, 'r', encoding=enc) as f:
                f.read(100)
            encoding = enc
            break
        except (UnicodeDecodeError, UnicodeError):
            continue

    if not encoding:
        encoding = 'gbk'

    with open(filepath, 'r', encoding=encoding) as f:
        in_data = False
        for line in f:
            line = line.strip()
            if line.startswith('交易时间,交易分类'):
                in_data = True
                continue
            if not in_data or not line:
                continue

            parts = line.split(',')
            if len(parts) < 9:
                continue

            date_str = parts[0].strip()
            counterparty = parts[2].strip()
            description = parts[4].strip()
            direction = parts[5].strip()
            amount_str = parts[6].strip()
            payment = parts[7].strip()
            status = parts[8].strip()

            if status != '交易成功':
                continue

            try:
                dt = datetime.strptime(date_str, '%Y-%m-%d %H:%M:%S')
            except ValueError:
                continue

            if target_months and dt.month not in target_months:
                continue

            desc = counterparty
            if description and description != counterparty:
                desc = f"{counterparty} - {description}"

            try:
                amount = float(amount_str)
            except ValueError:
                continue

            records.append({
                'date': date_str,
                'type': '支出' if direction == '支出' else '收入',
                'amount': amount,
                'description': desc,
                'platform': '支付宝',
                'category': '',
                'note': payment if payment else ''
            })

    return records


def parse_wechat_xlsx(filepath, target_months=None):
    """解析微信 XLSX，返回记录列表"""
    records = []
    wb = openpyxl.load_workbook(filepath, read_only=True)
    ws = wb.active

    # 找到表头行（含"交易时间"的行）
    header_row = None
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=25, values_only=True), start=1):
        if row[0] and '交易时间' in str(row[0]):
            header_row = row_idx
            break

    if not header_row:
        wb.close()
        return records

    data_start = header_row + 1

    for row in ws.iter_rows(min_row=data_start, max_row=ws.max_row, values_only=True):
        if not row[0] or not isinstance(row[0], datetime):
            continue

        dt = row[0]
        if target_months and dt.month not in target_months:
            continue

        direction = str(row[4]).strip() if row[4] else ''
        if direction not in ('支出', '收入'):
            continue

        status = str(row[7]).strip() if row[7] else ''
        if '成功' not in status and '已收' not in status and '已转账' not in status:
            continue

        counterparty = str(row[2]).strip() if row[2] else ''
        product = str(row[3]).strip() if row[3] else ''
        amount = float(row[5]) if row[5] else 0
        payment = str(row[6]).strip() if row[6] else ''

        desc = counterparty
        if product and not product.startswith(counterparty):
            desc = f"{counterparty} - {product}"

        records.append({
            'date': dt.strftime('%Y-%m-%d %H:%M:%S'),
            'type': direction,
            'amount': amount,
            'description': desc,
            'platform': '微信',
            'category': '',
            'note': payment if payment else ''
        })

    wb.close()
    return records


def group_by_month(records):
    """按月份分组"""
    groups = {}
    for r in records:
        month_key = r['date'][:7]  # "2026-05"
        groups.setdefault(month_key, []).append(r)
    return groups


def main():
    parser = argparse.ArgumentParser(description='解析支付宝/微信账单文件，输出按月分组的 JSON')
    parser.add_argument('files', nargs='+', help='账单文件路径（CSV 或 XLSX）')
    parser.add_argument('--months', type=str, default='',
                        help='过滤月份，逗号分隔，如 "5,6"。不传则保留所有月份')
    args = parser.parse_args()

    # 解析目标月份
    target_months = None
    if args.months:
        target_months = set(int(m.strip()) for m in args.months.split(','))

    all_records = []

    for filepath in args.files:
        if not os.path.exists(filepath):
            print(f"警告：文件不存在 {filepath}", file=sys.stderr)
            continue

        file_type = detect_file_type(filepath)
        if file_type == 'alipay_csv':
            records = parse_alipay_csv(filepath, target_months)
            platform = '支付宝'
        elif file_type == 'wechat_xlsx':
            records = parse_wechat_xlsx(filepath, target_months)
            platform = '微信'
        else:
            print(f"警告：无法识别文件类型 {filepath}", file=sys.stderr)
            continue

        all_records.extend(records)
        print(f"{platform} {os.path.basename(filepath)}: {len(records)} 条", file=sys.stderr)

    # 按月分组
    monthly = group_by_month(all_records)

    # 输出
    result = {
        'total': len(all_records),
        'months': monthly
    }

    print(json.dumps(result, ensure_ascii=False, indent=2))


if __name__ == '__main__':
    main()
