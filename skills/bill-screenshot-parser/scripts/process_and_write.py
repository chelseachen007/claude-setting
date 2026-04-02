#!/usr/bin/env python3
"""
账单数据处理与 Excel 写入
输入：JSON 交易记录 + 工作目录 + 年份
输出：写入 Excel 并返回统计 JSON
"""

import json
import os
import shutil
import sys
from datetime import datetime

from openpyxl import load_workbook

TEMPLATE_PATH = '/Users/chenjie/.claude/skills/bill-screenshot-parser/template.xlsx'

CATEGORIES = ['饮食', '食材', '住房', '咖啡', '通讯', '服饰', '交通',
              '彩妆', '社交', '医疗', '学习', '娱乐', '健身', '日用', '旅行', '其他']

CATEGORY_KEYWORDS = {
    '饮食': ['外卖', '餐厅', '美团', '饿了么', '肯德基', '麦当劳', '面馆', '海底捞',
            '快餐', '美食', '饮料', '零食', '小吃', '餐饮', '嵊州', '奶茶', '烤肉',
            '火锅', '烧烤', '甜品', '面包', '烘焙', '饭', '粉', '粥'],
    '食材': ['盒马', '超市', '菜场', '生鲜', '菜市场'],
    '住房': ['房租', '物业', '水电', '燃气', '宽带', '水费', '电费'],
    '咖啡': ['咖啡', 'Coffee', '星巴克', '瑞幸', 'Manner'],
    '通讯': ['话费', '充值', '移动', '联通', '电信', '通讯'],
    '服饰': ['服装', '衣服', '鞋', 'ZARA', '优衣库'],
    '交通': ['打车', '滴滴', '地铁', '公交', '加油', '停车', '高德', '充电',
            '出行', '中石化', '石化', '交通', '交通出行'],
    '彩妆': ['化妆品', '口红', '粉底', '护肤', '美容', '屈臣氏', '丝芙兰'],
    '社交': ['红包', '转账', '请客', '聚餐', '礼物'],
    '医疗': ['医院', '药店', '诊所', '体检', '挂号', '药品'],
    '学习': ['课程', '书籍', '培训', '知乎', '教育', '网课', '订阅'],
    '娱乐': ['电影', '游戏', '音乐', '视频', '爱奇艺', '腾讯视频', 'B站',
            'BUFF', 'steam', '腾讯天游', '娱乐'],
    '健身': ['健身', '运动', '瑜伽', '游泳', '体育馆'],
    '日用': ['便利', '购', '日用', '洗护', '纸巾', '京东', '拼多多', '淘宝'],
    '旅行': ['机票', '酒店', '旅游', '景点', '民宿', '火车票'],
}

# 排除关键词
EXCLUDE_KEYWORDS = [
    '余额宝', '收益发放', '基金分红', '理财', '余额宝收益',
    '花呗还款', '您于', '使用"先', '商户单号',
    '转入余额宝', '从余额宝', '自动充值',
]


def infer_category(description):
    desc = str(description)
    for cat, keywords in CATEGORY_KEYWORDS.items():
        if any(kw in desc for kw in keywords):
            return cat
    if any(kw in desc for kw in ['公益', '捐赠']):
        return '其他'
    return '其他'


def fix_year(date_str, year):
    """修正 OCR 常见的年份识别错误"""
    if not date_str:
        return date_str
    # 替换各种错误年份为目标年份
    import re
    return re.sub(r'\b(20[0-9]{2})', str(year), date_str, count=1)


def should_exclude(record):
    """判断是否应排除该记录"""
    desc = str(record.get('description', ''))
    note = str(record.get('note', ''))
    text = desc + note

    for kw in EXCLUDE_KEYWORDS:
        if kw in text:
            return True

    # 金额为 0 的记录排除
    if record.get('amount', 0) == 0:
        return True

    # 社交类大额转账排除
    if record.get('category') == '社交' and abs(record.get('amount', 0)) >= 5000:
        return True

    return False


def _strip_refund_prefix(desc):
    """去掉退款描述中的前缀，提取原始商户名"""
    import re
    # 匹配 "退款-xxx" 或 "退款xxx" 或 "已退款(xxx)"
    m = re.match(r'^(?:退款[-]?|已退款[（(])', str(desc))
    if m:
        stripped = str(desc)[m.end():]
        # 去掉尾部的括号和金额，如 "已退款(0.68)" → 只保留商户名部分
        stripped = re.sub(r'[（(][\d.]+[）)]$', '', stripped).strip()
        return stripped
    return str(desc)


def merge_dual_amounts(records):
    """合并支付宝双金额记录（消费+退款），按描述匹配而非相邻位置"""
    # 先按日期排序
    records = sorted(records, key=lambda x: x.get('date', ''))

    # 找出所有退款记录的索引
    refund_indices = set()
    for i, r in enumerate(records):
        desc = str(r.get('description', ''))
        if ('退款' in desc and r.get('type') == '收入' and r.get('amount', 0) > 0):
            refund_indices.add(i)

    # 为每个退款记录找匹配的消费记录
    matched_refunds = set()  # 已匹配的退款索引
    matched_origins = set()  # 已匹配的消费索引
    merge_pairs = []  # (origin_idx, refund_idx)

    for ri in refund_indices:
        refund = records[ri]
        refund_desc_stripped = _strip_refund_prefix(refund['description'])

        for oi, origin in enumerate(records):
            if oi in matched_origins or oi == ri:
                continue
            if origin.get('type') != '支出' or origin.get('amount', 0) <= 0:
                continue
            # 同一天 + 描述匹配
            if origin['date'][:10] != refund['date'][:10]:
                continue
            origin_desc = str(origin.get('description', ''))
            # 退款描述去掉前缀后包含原描述，或原描述包含退款描述
            if (refund_desc_stripped and origin_desc
                    and (refund_desc_stripped in origin_desc
                         or origin_desc in refund_desc_stripped)):
                merge_pairs.append((oi, ri))
                matched_origins.add(oi)
                matched_refunds.add(ri)
                break

    # 构建合并结果
    merged = []
    for i, r in enumerate(records):
        if i in matched_refunds:
            continue  # 退款记录已被合并，跳过
        if i in matched_origins:
            # 找到对应的退款记录
            for oi, ri in merge_pairs:
                if oi == i:
                    refund = records[ri]
                    orig_amount = abs(float(r.get('amount', 0)))
                    refund_amount = abs(float(refund.get('amount', 0)))
                    net_amount = round(orig_amount - refund_amount, 2)

                    r['amount'] = net_amount
                    r['type'] = '支出'
                    r['note'] = f"退款 {refund_amount:.2f} 元"
                    break
        merged.append(r)

    return merged


def normalize_amount(record):
    """标准化金额：负数=支出，正数=收入"""
    amt = record.get('amount', 0)
    try:
        amt = float(amt)
    except (ValueError, TypeError):
        amt = 0

    # 字符串金额处理
    if isinstance(record.get('amount'), str):
        s = record['amount'].strip()
        if s.startswith('-'):
            amt = abs(float(s.replace('-', '')))
            record['type'] = '支出'
        elif s.startswith('+'):
            amt = float(s.replace('+', ''))
            record['type'] = '收入'

    record['amount'] = abs(amt)

    if 'type' not in record or not record['type']:
        record['type'] = '支出' if amt < 0 else '收入'

    return record


def get_or_create_bill(work_dir):
    """获取或创建当月账单文件"""
    today = datetime.now()
    month_name = f"{today.year}年{today.month}月账单.xlsx"
    bill_path = os.path.join(work_dir, month_name)

    if os.path.exists(bill_path):
        return bill_path, 'append'
    else:
        shutil.copy(TEMPLATE_PATH, bill_path)
        return bill_path, 'create'


def write_to_excel(bill_path, records):
    """追加记录到 Excel（去重 + 排序）"""
    wb = load_workbook(bill_path)
    ws = wb['明细']

    # 读取已有记录用于去重
    existing_keys = set()
    last_row = ws.max_row
    for row in range(2, last_row + 1):
        date_val = ws.cell(row=row, column=1).value
        amount_val = ws.cell(row=row, column=3).value
        desc_val = ws.cell(row=row, column=5).value
        if date_val and amount_val:
            existing_keys.add((str(date_val), float(str(amount_val)) if amount_val else 0, str(desc_val)))

    # 按 (date, amount, description) 去重
    added = 0
    for r in sorted(records, key=lambda x: x.get('date', '')):
        key = (r['date'], r['amount'], r['description'])
        if key in existing_keys:
            continue
        existing_keys.add(key)

        last_row += 1
        ws.cell(row=last_row, column=1, value=r['date'])
        ws.cell(row=last_row, column=2, value=r['type'])
        ws.cell(row=last_row, column=3, value=r['amount'])
        ws.cell(row=last_row, column=4, value=r['category'])
        ws.cell(row=last_row, column=5, value=r['description'])
        ws.cell(row=last_row, column=6, value=r.get('platform', ''))
        ws.cell(row=last_row, column=7, value=r.get('note', ''))
        added += 1

    wb.save(bill_path)
    return added


def main():
    import argparse

    parser = argparse.ArgumentParser(description='处理账单数据并写入 Excel')
    parser.add_argument('--data', required=True, help='JSON 交易记录（数组或含 records 字段的对象）')
    parser.add_argument('--work-dir', required=True, help='工作目录（账单文件存放位置）')
    parser.add_argument('--year', type=int, default=datetime.now().year, help='目标年份')
    args = parser.parse_args()

    # 解析输入数据
    raw = json.loads(args.data)
    if isinstance(raw, dict) and 'records' in raw:
        records = raw['records']
    elif isinstance(raw, list):
        records = raw
    else:
        print(json.dumps({"error": "无效数据格式"}, ensure_ascii=False))
        sys.exit(1)

    # 数据处理流水线
    processed = []
    for r in records:
        # 1. 年份修正
        if r.get('date'):
            r['date'] = fix_year(r['date'], args.year)

        # 2. 标准化金额
        r = normalize_amount(r)

        # 3. 确保平台字段
        if not r.get('platform'):
            r['platform'] = '支付宝'

        # 4. 推断类别
        if not r.get('category'):
            r['category'] = infer_category(r.get('description', ''))

        # 4. 排除检查
        if should_exclude(r):
            continue

        processed.append(r)

    # 5. 双金额合并
    processed = merge_dual_amounts(processed)

    # 5.5 合并后过滤金额为 0 的记录
    processed = [r for r in processed if r.get('amount', 0) != 0]

    # 6. 获取/创建账单文件
    bill_path, mode = get_or_create_bill(args.work_dir)

    # 7. 写入 Excel（含去重 + 排序）
    added = write_to_excel(bill_path, processed)

    # 统计
    expense_total = sum(r['amount'] for r in processed if r['type'] == '支出')
    income_total = sum(r['amount'] for r in processed if r['type'] == '收入')
    cat_stats = {}
    for r in processed:
        if r['type'] == '支出':
            cat_stats[r['category']] = cat_stats.get(r['category'], 0) + r['amount']

    result = {
        "bill_path": bill_path,
        "mode": mode,
        "total_input": len(records),
        "after_filter": len(processed),
        "added": added,
        "skipped_duplicates": len(processed) - added,
        "expense_total": round(expense_total, 2),
        "income_total": round(income_total, 2),
        "net_expense": round(expense_total - income_total, 2),
        "categories": {k: round(v, 2) for k, v in sorted(cat_stats.items(), key=lambda x: -x[1])},
        "records": processed
    }

    print(json.dumps(result, ensure_ascii=False, indent=2))


if __name__ == '__main__':
    main()
